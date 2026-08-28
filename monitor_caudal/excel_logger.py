"""Registro de cada lectura de Flow (y el valor ingresado vigente) en un Excel."""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = ["Fecha y hora", "Flow (Nm3/h)", "Valor ingresado"]


class ExcelLogger:
    def __init__(self, path):
        self.path = Path(path)
        if self.path.exists():
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Datos"
            self.ws.append(HEADERS)
            self.wb.save(self.path)

    def log(self, flow_value, setpoint_value):
        row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), flow_value, setpoint_value]
        self.ws.append(row)
        self.wb.save(self.path)
